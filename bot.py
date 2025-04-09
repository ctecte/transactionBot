# print ("hello world")
# message handler for /- or something like that
# a regex to determine the components
# sample input /- $1.00 Wok hey x2
# / followed by + or - (not sure if i need to implement + logic)
# $ followed by a double value
# Name of item purchased
# Look for x%d, if x%d present, input that as quantity
# else set quantity to 1
# write this data to excel 

import telebot
import logging
import re
import time
import signal
import sys
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers
from dotenv import load_dotenv
import os


load_dotenv("variables.env") # Replace with own bots token from botfather
bot_token = os.getenv("bot_token")

bot = telebot.TeleBot(bot_token)
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

def signal_handler(sig, frame):
    print("Graceful shutdown")
    bot.stop_polling()  # Stop bot
    sys.exit(0)
 
# Handle SIGINT (CTRL+C)
signal.signal(signal.SIGINT, signal_handler)

MONTHS = ['January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December']

def create_new_excel_file(year):
    wb = Workbook()
    # Remove the default sheet that is automatically created.
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    for month in MONTHS:
        ws = wb.create_sheet(title=month)
        # Set header row
        ws.append(['Date', 'Name', 'Cost', 'Quantity', 'Item Type'])
        # Optionally, apply additional formatting here if needed
        # For example: setting column widths or styles
        for row in ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                cell.number_format = numbers.FORMAT_DATE_DDMMYY
        
    file_name = f"transactions{year}.xlsx"
    wb.save(file_name)
    return file_name

def add_transaction_to_excel(date, cost, name, quantity, item_type):
    # locate the current year's excel sheet
    current_year = datetime.today().strftime('%Y')
    current_month = datetime.today().strftime('%B')

    path_to_excel_sheet = f"transactions{current_year}.xlsx"

    print(path_to_excel_sheet)
    print(current_month)
    # print(today)
    new_data = [[date, name, cost, quantity,item_type]]
    try:
        loadedExcel = load_workbook(path_to_excel_sheet)
    except FileNotFoundError as e:
        loadedExcel = load_workbook(create_new_excel_file(current_year))
    current_sheet = loadedExcel[f'{current_month}']

    for row in new_data:
        current_sheet.append(row)
        new_row_index = current_sheet.max_row
        # Set custom date format for the first column (assuming the date is in column 1)
        date_cell = current_sheet.cell(row=new_row_index, column=1)
        date_cell.number_format = 'DD-MM-YYYY'
        
        # Set custom cost format for the cost column (assuming cost is in column 3)
        cost_cell = current_sheet.cell(row=new_row_index, column=3)
        cost_cell.number_format = '#,##0.00'
    
    
    
    loadedExcel.save(path_to_excel_sheet)

def todays_transactions():
    current_year = datetime.today().strftime('%Y')
    current_month = datetime.today().strftime('%B')

    path_to_excel_sheet = f"transactions{current_year}.xlsx"
    try:
        # Load the Excel file
        loadedExcel = load_workbook(path_to_excel_sheet)  # Might throw FileNotFoundError

        # Get the current month's sheet
        if current_month in loadedExcel.sheetnames:
            current_sheet = current_month
        else:
            print(f"Sheet '{current_month}' not found in {path_to_excel_sheet}.")
            return None

        df = pd.read_excel(path_to_excel_sheet, sheet_name=current_sheet)
        return df
        
    except FileNotFoundError:
        print(f"File '{path_to_excel_sheet}' not found.")
        return None

@bot.message_handler(commands=['month'])
def get_months_transactions(message):
    df = todays_transactions()  # Call the function to get the DataFrame
    if df is None or df.empty:
        bot.reply_to(message, "No transactions found for the month.")
        return
    
    # Ensure correct column order and remove unnecessary indexes
    df = df[['Date', 'Name', 'Cost', 'Quantity', 'Item Type']].dropna()

    # Convert the Date column to a datetime and then format it as dd-mm-yyyy
    df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%d-%m-%Y')
    # Convert cost to 2dp for consistent output
    df['Cost'] = df['Cost'].astype(float).round(2)  # Keep as float but rounded

    # Build mobile-friendly message
    response = ["📆 This month's Transactions:\n"]
    try :
        for date, group in df.groupby("Date"):
            # response.append(f"{date[:15].ljust(2)} $  Qty  Type")
            response.append(f"{date}")
            for idx, row in df.iterrows():
                if (date == f"{row['Date']}") :
                    response.append(
                        f"▫️ "
                        f"{row['Name'][:15].ljust(15)} "
                        f"${row['Cost']} "
                        f"(x{row['Quantity']}) "
                        f"{row['Item Type']}"
                    )
    except Exception as e:
        bot.reply_to(message, f"Error: {str(e)}")
        return
    
    # Format with smart wrapping
    # final_text = "```\n" + "\n".join(response) + "\n```"
    # bot.reply_to(message, final_text, parse_mode="MarkdownV2")
    bot.reply_to(message, "\n".join(response))
    


@bot.message_handler(commands=['year'])
def get_all_transactions(message):

    current_year = datetime.today().strftime('%Y')
    path_to_excel_sheet = f"transactions{current_year}.xlsx"

    isEmpty = True
    response = ["📆 This year's Transactions:\n"]

    for month in MONTHS:
        try: 
            df = pd.read_excel(path_to_excel_sheet, sheet_name=month)
        except:
            continue

        if df is None :
            continue
        
        # Ensure correct column order and remove unnecessary indexes
        df = df[['Date', 'Name', 'Cost', 'Quantity', 'Item Type']].dropna()

        if df.empty: 
            continue
        else:
            isEmpty = False


        df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%d-%m-%Y')
        # Convert cost to 2dp
        df['Cost'] = df['Cost'].astype(float).round(2)  
        response.append(f"**{month}**")
        # Build mobile-friendly message
        try :
            for date, group in df.groupby("Date"):
                response.append(f"{date}")
                for idx, row in df.iterrows():
                    if (date == f"{row['Date']}") :
                        response.append(
                            f"▫️ "
                            f"{row['Name'][:15].ljust(15)} "
                            f"${row['Cost']} "
                            f"(x{row['Quantity']}) "
                            f"{row['Item Type']}"
                        )
        except Exception as e:
            bot.reply_to(message, f"Error: {str(e)}")
            return
    
    if isEmpty:
        bot.reply_to(message, "No transactions found")
    else :
        # already set to markdown mode, not sure why tele is not bolding my months
        bot.reply_to(message, "\n".join(response), parse_mode= "Markdown")

@bot.message_handler(commands=['excel'])
def send_excel_file(message):
    current_year = datetime.today().strftime('%Y')
    path_to_excel_sheet = f"transactions{current_year}.xlsx"

    try:
        # Send the Excel file to the user
        with open(path_to_excel_sheet, 'rb') as file:
            bot.send_document(message.chat.id, file)
    except FileNotFoundError:
        bot.reply_to(message, "Excel not created yet. /start to start adding and creating your own sheet!")
    except Exception as e:
        bot.reply_to(message, f"An error occurred: {str(e)}")


@bot.message_handler(commands=["start", "help"])
def send_welcome(message): 
    reply = (
        "Enter your transaction in the following format:\n"
        "$amt Name xQuantity (if any)\n"
        "Eg: $10.00 Wok Hey Hotel81 x2"
    )
    bot.reply_to(message, reply)

def get_item_type(message_text):
    if "food" in message_text:
        return "Food"
    elif "drink" in message_text:
        return "Drink"
    elif "grocery" in message_text:
        return "Groceries"
    elif "item" in message_text:
        return "Item"
    else:
        return "Others"
    
@bot.message_handler(commands=["deleteLast"])
def delete_last_transaction(message):
    try: 
        current_year = datetime.today().strftime('%Y')
        current_month = datetime.today().strftime('%B')
        path_to_excel_sheet = f"transactions{current_year}.xlsx"

        wb = load_workbook(path_to_excel_sheet)
        sheet = wb[current_month]

        if (sheet.max_row <= 1): 
            # month is empty, last transaction is in a previous month
            foundLastTransaction = False
            for month in reversed(MONTHS):
                sheet = wb[month]
                if (sheet.max_row > 1):
                    foundLastTransaction = True
                    break
        else: 
            foundLastTransaction = True

        if (not foundLastTransaction):
            bot.reply_to(message, "There is no previous transactions in the excel")
            return
        
        deleted_row = ""
        for cell in sheet[sheet.max_row]:
            val = cell.value
            if isinstance(val, datetime):
                deleted_row += val.strftime("%d-%m-%y") + " "
            else:
                deleted_row += str(val) + " "
        
        try: 
            sheet.delete_rows(sheet.max_row, 1)
            wb.save(path_to_excel_sheet)
            bot.reply_to(message, f"successfully deleted row\n{deleted_row}")
            return
        except Exception as e:
            bot.reply_to(message, f"Failed to delete. Error: {str(e)}")
            return


    except Exception as e:
        bot.reply_to(message, f"Error: {str(e)}")
        return
    
@bot.message_handler(commands=["backdate"])
def backdate_transaction(message):
    try:
        # transaction_text = message.text.replace("/spend", "", 1).strip()
        item_type = get_item_type(message.text.lower())
        # remove the /backdate from the message
        command = message.text.split(" ", 1)[0]
        transaction_text = message.text.replace(command, "", 1).strip()
        # regex expects input of "200325 $10 wokhey x2"
        match = re.match(r"(\d{6})\s+\$(\d+(?:\.\d{1,2})?)\s+(.+?)(?:\s+x(\d+))?$", transaction_text)

        if match:
            date = datetime.strptime(match.group(1), "%d%m%y").date()
            cost = float(match.group(2))  # Convert cost to float
            name = match.group(3)  # Extract name string
            # print(name)
            if match.group(4):
                quantity = int(match.group(4))  # Convert quantity to int
            else: 
                quantity = 1
            logger.info("parsed successfully")
            # bot.reply_to(message, "successfully parsed")
            logger.info({"cost": cost, "name": name, "quantity": quantity})
            today = datetime.today().date()
            if (date > today):
                bot.reply_to(message, "Cannot backdate to the future! Transaction not added")
            else: 
                add_transaction_to_excel(date, cost, name, quantity, item_type)
                bot.reply_to(message, "Transaction added")
        else:
            bot.reply_to(message, "invalid format, regex expecting\n/backdate DDMMYY $XX.XX name of food xQty food/drink/grocery/item")

    except FileNotFoundError as e:
        bot.reply_to(message, "Transaction excel sheet not found. Please check path_to_excel_sheet variable")
        logger.error(f"Error: {str(e)}")

    except Exception as e: 
        logger.error(f"Error: {str(e)}")
        bot.reply_to(message, f"Error: {str(e)}")
    

@bot.message_handler(commands=["food", "drink", "item", "grocery"])
def record_transaction(message):
    try:
       
        item_type = get_item_type(message.text.lower())
        # remove the /whatever from the message
        command = message.text.split(" ", 1)[0]
        transaction_text = message.text.replace(command, "", 1).strip()
        match = re.match(r"\$(\d+(?:\.\d{2})?)\s+(.+?)(?:\s+x(\d+))?$", transaction_text)
        logger.info("regexing..")

        if match:
            cost = float(match.group(1))  # Convert cost to float
            name = match.group(2)  # Extract name string
            # print(name)
            if match.group(3):
                quantity = int(match.group(3))  # Convert quantity to int
            else: 
                quantity = 1
            logger.info("parsed successfully")
            # bot.reply_to(message, "successfully parsed")
            logger.info({"cost": cost, "name": name, "quantity": quantity})
            today = datetime.today().date()
            add_transaction_to_excel(today ,cost, name, quantity, item_type)
            bot.reply_to(message, "Transaction added")
        else:
            bot.reply_to(message, "invalid format, please re-enter transaction. /start for template")

    except FileNotFoundError as e:
        bot.reply_to(message, "Transaction excel sheet not found. Please check path_to_excel_sheet variable")
        logger.error(f"Error: {str(e)}")

    except Exception as e: 
        logger.error(f"Error: {str(e)}")
        bot.reply_to(message, f"Error: {str(e)}")

logger.info("bot is now running")

if __name__ == "__main__":
    try:
        bot.polling(none_stop=True, interval=2, timeout=20)
    except Exception as e:
        print(f"Error: {e}")
        time.sleep(15)




